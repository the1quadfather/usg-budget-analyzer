"""
parsing/xlsx_ingest.py

Parses the official DoD Comptroller R-1 spreadsheet (r1_display.xlsx) into the
same normalized schema as the PDF-based r1_parser. These files exist for
FY2012-FY2027 (config.XLSX_FIRST_FY) and are authoritative, so no OCR or
text-layout heuristics are needed for those years.

Ground-truth structure (verified against FY2027 r1_display.xlsx):
  - Sheet "Exhibit R-1" holds every column; per-scenario sheets duplicate it
  - Row 1: "Total of Displayed Rows" totals; row 2: headers; data from row 3
  - Base columns: Account, Account Title, Organization, Budget Activity,
    Budget Activity Title, Line Number, PE/BLI, PE/BLI Title, Include In TOA,
    Classification
  - Amount columns are per funding scenario, e.g. FY2027 file:
      FY 2025 Actuals | FY 2025 Reconciliation | FY 2025 Total
      FY 2026 Discretionary Enacted | FY 2026 PL 119-21 Spend Plan | FY 2026 Total
      FY 2027 Discretionary Request | FY 2027 Mandatory Request | FY 2027 Total
    Older files use Base / OCO / Total splits instead.
  - Dollars in THOUSANDS, stored as strings ('' for null)
  - Organization carries Defense-Wide sub-agency codes (DARPA, MDA, SOCOM...)
    that the PDF R-1 does not expose - preserved in an extra column
  - Non-RDT&E accounts (OIG, Chemical Agents, Golden Dome fund...) appear in
    the workbook; excluded by default to stay consistent with the PDF-parsed
    historical series (r1_parser skips them too)

Output: r1_parser.OUTPUT_COLUMNS with extraction_method="xlsx", plus extras:
    organization          str    sub-agency code (A, N, F, DARPA, MDA, ...)
    py_mandatory_amount   float  reconciliation/mandatory $K for PY, if any
    cy_mandatory_amount   float  reconciliation/mandatory $K for CY, if any
    by_mandatory_amount   float  reconciliation/mandatory $K for BY, if any

py/cy/by_amount carry the DISCRETIONARY stream (actuals / enacted / request)
so trend lines stay comparable with pre-reconciliation years; mandatory money
(P.L. 119-21 etc.) rides in the *_mandatory_amount columns.

Usage:
    parser = R1XlsxParser()
    df = parser.parse(Path("data/raw/comptroller/2027/rdtee/fy2027_r1.xlsx"))

CLI:
    python parsing/xlsx_ingest.py --file data/raw/comptroller/2027/rdtee/fy2027_r1.xlsx
"""

import logging
import re
from pathlib import Path

import openpyxl
import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
import config

try:
    from parsing.r1_parser import BA_MAP, OUTPUT_COLUMNS, NativeR1Parser, R1Parser
except ImportError:
    from r1_parser import BA_MAP, OUTPUT_COLUMNS, NativeR1Parser, R1Parser

logger = logging.getLogger(__name__)


# Accounts to keep - RDT&E appropriations and OT&E, matching the PDF parser's
# appropriation filter so xlsx- and pdf-derived years form one series.
RDTEE_ACCOUNT_RE = re.compile(
    r"research,\s*development,\s*test.*eval|operational\s+test\s+and\s+evaluation",
    re.IGNORECASE,
)

# Amount-column headers look like "FY 2026 Discretionary Enacted".
FY_COL_RE = re.compile(r"^FY\s*(\d{4})\s*(.*)$", re.IGNORECASE)

# Qualifier buckets. Primary = the discretionary series; mandatory = the
# reconciliation stream Congress added outside discretionary caps.
MANDATORY_QUALIFIER_RE = re.compile(r"reconciliation|mandatory|pl\s*119", re.IGNORECASE)
PRIMARY_QUALIFIER_PRIORITY = [
    "actuals",
    "discretionary enacted",
    "enacted",
    "discretionary request",
    "base",
    "request",
    "total",   # fallback only - includes OCO/mandatory when no split exists
]

EXTRA_COLUMNS = [
    "organization",
    "py_mandatory_amount", "cy_mandatory_amount", "by_mandatory_amount",
]


def _to_amount(value) -> float | None:
    """Spreadsheet amounts are strings in $K; '' means no funding."""
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


class R1XlsxParser:
    """Parses one r1_display.xlsx into the shared R-1 schema."""

    def parse(
        self,
        xlsx_path: Path,
        fiscal_year: int | None = None,
        include_non_rdtee: bool = False,
    ) -> pd.DataFrame:
        xlsx_path = Path(xlsx_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(xlsx_path)
        if fiscal_year is None:
            fiscal_year = R1Parser._infer_fy(xlsx_path)

        logger.info(f"Parsing {xlsx_path.name} (FY{fiscal_year}, xlsx)")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            ws, headers, header_idx = self._find_exhibit_sheet(wb)
            col = {h: i for i, h in enumerate(headers) if h}
            fy_cols = self._classify_amount_columns(headers, fiscal_year)

            records = []
            skipped_accounts: set[str] = set()
            for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
                rec = self._parse_row(
                    row, col, fy_cols, fiscal_year, xlsx_path.name,
                    include_non_rdtee, skipped_accounts,
                )
                if rec:
                    records.append(rec)
        finally:
            wb.close()

        if skipped_accounts:
            logger.info(f"  Skipped non-RDT&E accounts: {sorted(skipped_accounts)}")
        if not records:
            logger.warning(f"  No records extracted from {xlsx_path.name}")
            return pd.DataFrame(columns=OUTPUT_COLUMNS + EXTRA_COLUMNS)

        df = R1Parser._normalise(pd.DataFrame(records))
        logger.info(f"  -> {len(df):,} clean PE records")
        return df

    # ── Workbook structure ────────────────────────────────────────────────────

    @staticmethod
    def _find_exhibit_sheet(wb):
        """
        Return (worksheet, headers, header_row_index) for the main exhibit
        sheet. Prefers a sheet named like "Exhibit R-1"; otherwise takes the
        first sheet whose top rows contain a PE/BLI header.
        """
        candidates = [n for n in wb.sheetnames if "exhibit" in n.lower()]
        candidates += [n for n in wb.sheetnames if n not in candidates]

        for name in candidates:
            ws = wb[name]
            for idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                headers = [str(v).strip() if v is not None else "" for v in row]
                if any(h.upper().startswith("PE/BLI") or "PROGRAM ELEMENT" in h.upper()
                       for h in headers):
                    if "Account" in headers:
                        return ws, headers, idx
        raise ValueError("No sheet with an R-1 header row (Account / PE/BLI) found")

    @staticmethod
    def _classify_amount_columns(
        headers: list[str], fiscal_year: int
    ) -> dict[int, dict[str, int]]:
        """
        Map each fiscal year mentioned in the headers to its column indices:
            {2026: {"primary": 12, "mandatory": 13}, ...}
        Primary follows PRIMARY_QUALIFIER_PRIORITY; mandatory catches the
        reconciliation/mandatory stream.
        """
        by_fy: dict[int, dict[str, int]] = {}
        candidates: dict[int, list[tuple[str, int]]] = {}
        for i, h in enumerate(headers):
            m = FY_COL_RE.match(h)
            if not m:
                continue
            fy = int(m.group(1))
            qualifier = m.group(2).strip().lower()
            candidates.setdefault(fy, []).append((qualifier, i))

        for fy, quals in candidates.items():
            entry: dict[str, int] = {}
            for qualifier, i in quals:
                if MANDATORY_QUALIFIER_RE.search(qualifier) and "mandatory" not in entry:
                    entry["mandatory"] = i
            for wanted in PRIMARY_QUALIFIER_PRIORITY:
                hit = next(
                    (i for qualifier, i in quals
                     if qualifier == wanted and i != entry.get("mandatory")),
                    None,
                )
                if hit is not None:
                    entry["primary"] = hit
                    break
            if "primary" not in entry:   # unrecognised qualifiers - take first
                entry["primary"] = quals[0][1]
            by_fy[fy] = entry

        missing = [fy for fy in (fiscal_year - 2, fiscal_year - 1, fiscal_year)
                   if fy not in by_fy]
        if missing:
            logger.warning(f"  Headers lack columns for FY{missing}")
        return by_fy

    # ── Row parsing ───────────────────────────────────────────────────────────

    def _parse_row(
        self,
        row: tuple,
        col: dict[str, int],
        fy_cols: dict[int, dict[str, int]],
        fiscal_year: int,
        source_file: str,
        include_non_rdtee: bool,
        skipped_accounts: set[str],
    ) -> dict | None:
        def cell(name: str):
            i = col.get(name)
            return row[i] if i is not None and i < len(row) else None

        pe_number = str(cell("PE/BLI") or "").strip()
        if not pe_number:
            return None

        account_title = str(cell("Account Title") or "").strip()
        if not include_non_rdtee and not RDTEE_ACCOUNT_RE.search(account_title):
            if account_title:
                skipped_accounts.add(account_title)
            return None

        # Title column header is verbose ("Program Element/Budget Line Item...")
        title_key = next(
            (h for h in col if "Program Element" in h or "BLI) T" in h), None
        )
        pe_title = str(cell(title_key) or "").strip() if title_key else ""

        raw_act = str(cell("Budget Activity") or "").strip()
        act_code = raw_act.zfill(2) if raw_act.isdigit() else raw_act
        ba_title = str(cell("Budget Activity Title") or "").strip()
        budget_activity = BA_MAP.get(
            act_code,
            f"BA{int(act_code)} - {ba_title}" if act_code.isdigit() else ba_title,
        )

        line_raw = str(cell("Line Number") or "").strip()
        classification = str(cell("Classification") or "").strip()

        def amount(fy: int, which: str):
            i = fy_cols.get(fy, {}).get(which)
            return _to_amount(row[i]) if i is not None and i < len(row) else None

        return {
            "fiscal_year": fiscal_year,
            "component": NativeR1Parser._normalise_component(account_title),
            "appropriation": account_title,
            "budget_activity": budget_activity,
            "pe_number": "" if pe_number.startswith("999") else pe_number,
            "pe_title": pe_title,
            "line_no": int(line_raw) if line_raw.isdigit() else None,
            "act_code": act_code,
            "is_classified": (
                pe_number.startswith("999")
                or classification not in ("", "U")
            ),
            "py_amount": amount(fiscal_year - 2, "primary"),
            "cy_amount": amount(fiscal_year - 1, "primary"),
            "by_amount": amount(fiscal_year, "primary"),
            "source_file": source_file,
            "extraction_method": "xlsx",
            "organization": str(cell("Organization") or "").strip(),
            "py_mandatory_amount": amount(fiscal_year - 2, "mandatory"),
            "cy_mandatory_amount": amount(fiscal_year - 1, "mandatory"),
            "by_mandatory_amount": amount(fiscal_year, "mandatory"),
        }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=config.LOG_LEVEL,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    ap = argparse.ArgumentParser(description="Parse an official R-1 xlsx exhibit.")
    ap.add_argument("--file", required=True, type=Path)
    ap.add_argument("--fy", type=int, default=None)
    ap.add_argument("--save", action="store_true",
                    help="Write r1_{FY}.parquet to the processed dir")
    args = ap.parse_args()

    df = R1XlsxParser().parse(args.file, fiscal_year=args.fy)
    print(df.head(20).to_string())
    print(f"\n{len(df):,} rows | components: {sorted(df['component'].unique())}")

    if args.save and not df.empty:
        fy = int(df["fiscal_year"].iloc[0])
        out = config.PROCESSED_DIR / f"r1_{fy}.parquet"
        df.to_parquet(out, index=False)
        print(f"Saved -> {out}")
